import io
import json
import os
import openpyxl
import pandas as pd
import streamlit as st

def get_target_sheet(wb, config_sheet_name):
    # 1. Direct match from JSON config
    if config_sheet_name in wb.sheetnames:
        return wb[config_sheet_name]

    # 2. Common localized sheet names used by Amazon
    common_amazon_sheets = ["Vorlage", "Modèle", "Modello", "Plantilla", "Template"]
    for name in common_amazon_sheets:
        for actual_name in wb.sheetnames:
            if name.lower() in actual_name.lower():
                return wb[actual_name]

    # 3. Fallback: Find sheet starting with "Template-", "Vorlage-", etc.
    for name in wb.sheetnames:
        if any(prefix in name for prefix in ["Template-", "Vorlage-", "Modèle-"]):
            return wb[name]

    # 4. Final safety fallback: Assume the last worksheet is the data tab
    return wb.worksheets[-1]

st.set_page_config(page_title="Excel Static Mapping Automation", page_icon="🚀", layout="wide")
st.title("🌍 Universal Excel Static Mapping Automation")

# Setup project relative directories
BASE_DIR = os.path.dirname(__file__)
CONFIG_DIR = os.path.join(BASE_DIR, "country_config")
SC_TEMPLATES_DIR = os.path.join(BASE_DIR, "SC_Templates")

# Validate repository integrity
if not os.path.exists(CONFIG_DIR) or not os.path.exists(SC_TEMPLATES_DIR):
    st.error("❌ Repository missing required folders: 'country_config' or 'SC_Templates'.")
    st.stop()

# ---------------------------------------------------------
# 1. Dynamic UI Navigation (Category > Subcategory > Country)
# ---------------------------------------------------------
st.sidebar.header("Step 1: Choose Profile")

# Discover top-level Categories (Folders in country_config)
categories = [
    d for d in os.listdir(CONFIG_DIR) if os.path.isdir(os.path.join(CONFIG_DIR, d))
]

if not categories:
    st.sidebar.error("⚠️ No category folders found inside `country_config/`.")
    st.stop()

selected_category = st.sidebar.selectbox("Select Main Category", categories)
category_path = os.path.join(CONFIG_DIR, selected_category)

# Discover Subcategories or Direct Profiles
subitems = os.listdir(category_path)
has_subcategories = any(os.path.isdir(os.path.join(category_path, s)) for s in subitems)

if has_subcategories:
    subcategories = [s for s in subitems if os.path.isdir(os.path.join(category_path, s))]
    selected_subcategory = st.sidebar.selectbox("Select Subcategory", subcategories)
    active_config_dir = os.path.join(category_path, selected_subcategory)
else:
    selected_subcategory = None
    active_config_dir = category_path

available_countries = [
    os.path.splitext(f)[0] for f in os.listdir(active_config_dir) if f.endswith(".json")
]

if not available_countries:
    st.sidebar.warning("⚠️ No country profiles found for this selection.")
    st.stop()

target_country = st.sidebar.selectbox("Select Target Country", available_countries)

# ---------------------------------------------------------
# 2. Metadata Extraction
# ---------------------------------------------------------
config_path = os.path.join(active_config_dir, f"{target_country}.json")
with open(config_path, "r", encoding="utf-8") as f:
    meta = json.load(f)

m_id = meta.get("m_id")
lang = meta.get("lang")
target_sheet = meta.get("sheet_name", "Template-WATCH")
sc_template_name = meta.get("sc_name_pattern")
vc_suffix = meta.get("vc_name_suffix", "")

st.sidebar.markdown("---")
st.sidebar.info(
    f"**Category:** {selected_category}" + (f" ({selected_subcategory})" if selected_subcategory else "") + "\n\n"
                                                                                                            f"**Marketplace ID:** {m_id}\n\n"
                                                                                                            f"**Language Tag:** {lang}\n\n"
                                                                                                            f"**Target Sheet:** `{target_sheet}`"
)

blank_sc_file = os.path.join(SC_TEMPLATES_DIR, sc_template_name)
if not os.path.exists(blank_sc_file):
    st.error(f"❌ Template missing: `{sc_template_name}` not found in `SC_Templates/`.")
    st.stop()

# ---------------------------------------------------------
# 3. Document Upload Gateway
# ---------------------------------------------------------
st.subheader("1. Source Document")
vc_file_upload = st.file_uploader(
    f"Upload Vendor Central File for {selected_category} ({target_country})",
    type=["xlsx", "xlsm"],
)

# ---------------------------------------------------------
# 4. Core Processing Engine
# ---------------------------------------------------------
if vc_file_upload:
    st.subheader("2. Execute Processing")

    if st.button("🚀 Run Automation Engine", type="primary"):
        output_prefix = selected_subcategory or selected_category
        output_filename = f"SC_{output_prefix}_Static_Mapped_{target_country}.xlsm"

        try:
            with st.spinner("Processing mapping schemas & localization..."):

                # A. Dynamic Static Mapping Resolution
                base_static_map = meta.get("base_mappings", {})
                custom_mappings = meta.get("additional_mappings", {})
                combined_map = {**base_static_map, **custom_mappings}

                current_static_map = {}
                for vc_col, sc_template_str in combined_map.items():
                    current_static_map[vc_col] = sc_template_str.format(m_id=m_id, lang=lang)

                # B. Static Default Values
                raw_static_values = meta.get("static_values", {})
                current_static_values = {
                    sc_template_str.format(m_id=m_id, lang=lang): val
                    for sc_template_str, val in raw_static_values.items()
                }

                # C. Read Vendor Central File
                xl = pd.ExcelFile(vc_file_upload)

                vc_sheet_to_use = None
                common_vc_sheets = [target_sheet, "Template-WATCH", "Template-BRACELET"]

                for name in common_vc_sheets:
                    for actual_name in xl.sheet_names:
                        if name.lower() in actual_name.lower():
                            vc_sheet_to_use = actual_name
                            break
                    if vc_sheet_to_use:
                        break

                # Fallback to the last sheet (often the data sheet in Amazon files) if none match
                if not vc_sheet_to_use:
                    vc_sheet_to_use = xl.sheet_names[-1]

                vc_df = pd.read_excel(vc_file_upload, sheet_name=vc_sheet_to_use, header=3)
                vc_df.columns = [str(c).strip() for c in vc_df.columns]

                # D. Build Dynamic Translation Table from Config
                cfg_trans = meta.get("translations", {})
                raw_translation_map = meta.get("dropdown_translation_keys", {})

                dropdown_translations = {}
                for sc_key_template, trans_group_key in raw_translation_map.items():
                    localized_sc_key = sc_key_template.format(m_id=m_id, lang=lang)
                    dropdown_translations[localized_sc_key] = cfg_trans.get(trans_group_key, {})

                # E. Load Target OpenPyXL Workbook
                wb = openpyxl.load_workbook(blank_sc_file, keep_vba=True)
                ws = get_target_sheet(wb, target_sheet)

                headers = [
                    str(cell.value).strip() if cell.value is not None else ""
                    for cell in ws[5]
                ]
                start_row = 7

                # F. Map Data Rows
                for idx, vc_row in vc_df.iterrows():
                    current_sku = str(vc_row.get("vendor_sku#1.value", "")).strip()
                    if isinstance(current_sku, pd.Series):
                        current_sku = str(current_sku.iloc[0]).strip()

                    if current_sku in ["ABC123", "nan", ""] or "REQUIRED" in current_sku.upper():
                        continue

                    # Map Columns
                    for vc_col, sc_col in current_static_map.items():
                        if vc_col in vc_df.columns:
                            cell_value = vc_row[vc_col]
                            if isinstance(cell_value, pd.Series):
                                cell_value = cell_value.iloc[0]
                            cell_value = "" if pd.isna(cell_value) else str(cell_value).strip()

                            if sc_col in headers:
                                col_idx = headers.index(sc_col) + 1
                                target_cell = ws.cell(row=start_row, column=col_idx)

                                # Translate Dropdowns if Applicable
                                if sc_col in dropdown_translations and cell_value != "":
                                    lookup_dict = dropdown_translations[sc_col]
                                    val_lower = cell_value.lower()
                                    matched_local_val = next(
                                        (v for k, v in lookup_dict.items() if k.lower() == val_lower),
                                        None
                                    )
                                    if matched_local_val:
                                        cell_value = matched_local_val

                                target_cell.value = None
                                if sc_col == "amzn1.volt.ca.product_id_value":
                                    target_cell.number_format = "@"
                                    target_cell.value = str(cell_value)
                                else:
                                    target_cell.value = cell_value

                    # Inject Static Values
                    for sc_col, static_value in current_static_values.items():
                        if sc_col in headers:
                            col_idx = headers.index(sc_col) + 1
                            ws.cell(row=start_row, column=col_idx).value = static_value

                    start_row += 1

                # G. Format Column Widths
                MAX_ALLOWED_WIDTH = 50
                for col in ws.columns:
                    max_len = 0
                    col_letter = col[0].column_letter
                    for row_idx in range(7, ws.max_row + 1):
                        val = ws.cell(row=row_idx, column=col[0].column).value
                        if val:
                            max_len = max(max_len, len(str(val)))
                    if max_len > 0:
                        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), MAX_ALLOWED_WIDTH)
                    else:
                        ws.column_dimensions[col_letter].width = 15

                wb.calculation.calcMode = "auto"

                # H. Output to Memory Stream
                excel_buffer = io.BytesIO()
                wb.save(excel_buffer)
                excel_buffer.seek(0)

            st.success("🎉 Mapping and Translation Completed Successfully!")

            st.download_button(
                label=f"📥 Download Processed File ({selected_category} - {target_country})",
                data=excel_buffer,
                file_name=output_filename,
                mime="application/vnd.ms-excel.sheet.macroEnabled.12",
            )

        except Exception as e:
            st.error(f"💥 Pipeline execution failure: {str(e)}")
else:
    st.info("💡 Drop your Vendor Central file above to begin transformation.")